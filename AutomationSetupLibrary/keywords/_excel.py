from robot.api import logger
from robot.libraries.BuiltIn import BuiltIn
from .keywordgroup import KeywordGroup

import openpyxl
import os

class _Excel(KeywordGroup):

    def __init__(self):
        self.builtin = BuiltIn()

    @property
    def curdir(self):
        return os.getcwd()

    @property
    def mass(self):
        return BuiltIn().get_variable_value("${MASS}")

    def open_excel_document(self, path, file, fixtures_path='\\resources\\fixtures\\'):
        """Opens an excel document to be read or written. It returns the workbook of the opened document.

        Arguments:
            - path: relative path where to find the file. If the file is not inside fixtures path, please,
            overwrite the fixtures_path parameter and use that one to concatenate with, or, just give it
            an empty string ''. Do not use '\\' if not overwrite fixtures_path. Follow the examples to see
            its usage.
            - file: file name with the excel extension.
            - (optional) fixtures_path: Parameter to overwrite the dafault value '\\resources\\fixtures\\'.
            You should use two '\\' with this parameter

        Examples:
        | Open excel document | reinf | R2040.xlsx |
        | Open excel document | reinf | R2040.xls | \\\\resources\\\\support\\\\ |

        =>

        | (robotframework suite variable) ${WORKBOOK} |
        """
        path_locator = self.curdir + fixtures_path + path
        wb = openpyxl.load_workbook(str(path_locator) + file)
        self.builtin.set_suite_variable("${WORKBOOK}", wb)
        return wb

    def get_excel_value_by_header_column_text(self, column_text, row_value, work_book=None):
        """Get the excel cell value giving an header column text with the respective row to return.

        Arguments:
            - column_text: column text to respective with the value you want to get. (String)
            - row_value: row_value respective to the value that should be returned. Rows start from 2,
            because 1 is stricted for the header. (int)
            - (optional) work_book: If given it will use the work_book of this parameter if non was found
            on the suite variable ${WORKBOOK}

        Examples:
        | get excel value by header column text | InvoiceDate | 3 |

        =>

        | value |
        """
        wb = self.builtin.get_variable_value("${WORKBOOK}", work_book)
        ws = wb.active
        i, r = 1, False
        for row in ws.values:
            for value in row:
                if column_text in value:
                    r = True
                    break
                i += 1
            if r: break
            i = 1
        return ws[str(chr(64 + int(i))) + str(row_value)]

    def change_excel_value_by_header_column_text(self, column_text, row_value, value, work_book=None):
        """Changes the excel cell value giving an header column text with the respective row to change.

        Arguments:
            - column_text: column text to respective with the value you want to change. (String)
            - row_value: row_value respective to the value that should be changed. Rows start from 2,
            because 1 is stricted for the header. (int)
            - value: Value that you want the cell to be filled with. (String, float, int, boolean)
            - (optional) work_book: If given it will use the work_book of this parameter if non was found
            on the suite variable ${WORKBOOK}.

        Examples:
        | change excel value by header column text | InvoiceDate | 3 | 01.01.2020 |

        =>

        | value |
        """
        wb = self.builtin.get_variable_value("${WORKBOOK}", work_book)
        ws = wb.active
        i, r, cell = 1, False, ""
        for row in ws.values:
            for v in row:
                if v is not None:
                    if column_text in v.replace(" ", ""):
                        r = True
                        break
                i += 1
            if r: break
            i = 1
        cell = str(chr(64 + int(i))) + str(row_value)
        ws[cell] = value
        ws[cell].number_format = "@"
        return ws[cell]

    def change_cell_on_excel(self, sheet, row, column, cell_value, work_book=None):
        """Changes the excel cell value giving all coordinates to the change.

        Arguments:
            - sheet: Sheet name where it should change its cell value. (String)
            - row_value: row_value respective to the value that should be changed. Normally rows starts from 2,
            because 1 is stricted for the header. (int)
            - column: column number to respective with the value you want to change. (int)
            - cell_value: Value that you want the cell to be filled with. (String, float, int, boolean)
            - (optional) work_book: If given it will use the work_book of this parameter if non was found
            on the suite variable ${WORKBOOK}.

        Examples:
        | change cell on excel | Sheet1 | 4 | 5 | 01.01.2020 |

        =>

        | value |
        """
        wb = self.builtin.get_variable_value("${WORKBOOK}", work_book)
        sheet = wb.get_sheet_by_name(sheet)
        cell =str(chr(64 + int(column))) + str(row)
        ws = wb.active
        ws[cell].number_format = "@"
        sheet.cell(row=int(row), column=int(column)).value = cell_value
        return ws[cell]

    def save_excel_changes(self, path, file_name, fixtures_path='\\resources\\fixtures\\', work_book=None):
        """Opens an excel document to be read or written. It returns the workbook of the opened document.

        Arguments:
            - path: relative path where to save the file. If the file is not inside fixtures path, please,
            overwrite the fixtures_path parameter and use that one to concatenate with, or, just give it
            an empty string ''. Do not use '\\' if not overwrite fixtures_path. Follow the examples to see
            its usage.
            - file_name: file name with the excel extension to be saved.
            - (optional) fixtures_path: Parameter to overwrite the dafault value '\\resources\\fixtures\\'.
            You should use two '\\' with this parameter
            - (optional) work_book: If given it will use the work_book of this parameter if non was found
            on the suite variable ${WORKBOOK}.

        Examples:
        | Save excel changes | reinf | R2040.xlsx |
        | Save excel changes | reinf | R2040.xls | \\\\resources\\\\support\\\\ |

        =>

        | (robotframework suite variable) ${WORKBOOK} |
        """
        wb = self.builtin.get_variable_value("${WORKBOOK}", work_book)
        path_to_save = self.curdir + fixtures_path + path + file_name
        wb.save(path_to_save)
        wb.close()

    def edit_excel_by_mass(self, file_path, file_name, save_path, save_file_name, start_row, mass):
        """Opens an excel document and edits it with a given mass.

        Arguments:
            - file_path: relative path where to save the file. If the file is not inside fixtures path, please,
            overwrite the fixtures_path parameter and use that one to concatenate with, or, just give it
            an empty string ''. Do not use '\\' if not overwrite fixtures_path. Follow the examples to see
            its usage.
            - file_name: file name with the excel extension to be saved.
            - save_path: Path to save the file. (String).
            - save_file_name: File name that should be used to save the excel document with the extension (String).
            - mass: Python dict to edit the excel. This method will use the key of the dict to find the column that
            matches the given key and assign its value to the excel cell. All keys with value should be inside a higher
            one that represents the excel line. e.g. ({"line_1": {"column_1": "value_1"}, "line_2": {"column_1": "value_2"}}). (dict)
            - start_row: Defines each row should start edit the excel file. (int)
            - (optional) fixtures_path: Parameter to overwrite the dafault value '\\resources\\fixtures\\'.
            You should use two '\\' with this parameter
            - (optional) work_book: If given it will use the work_book of this parameter if non was found
            on the suite variable ${WORKBOOK}.

        Examples:
        | edit excel by mass | \\\\resources\\fixtures\\reinf | R2040.xlsx | \\\\resources\\fixtures\\reinf\\edited | edited.xlsx | 1 | ${MASS["your_mass"]} |

        =>

        | (robotframework suite variable) ${WORKBOOK} |
        """
        self.open_excel_document(file_path, file_name)
        row = int(start_row)
        for j,l in self.mass[mass].items():
            for k,v in l.items():
                self.change_excel_value_by_header_column_text(k, row, v)
            row += 1
        self.save_excel_changes(save_path, save_file_name)

    def read_cell_on_excel(self, sheet, row, column, work_book=None):
        """Reads an cell on excell with its full coordinates

        Arguments:
            - sheet: Sheet name where it should change its cell value. (String)
            - row: row respective to the value that should be returned. Normally rows starts from 2,
            because 1 is stricted for the header. (int)
            - column: column number to respective with the value you want to get. (int)
            - (optional) work_book: If given it will use the work_book of this parameter if non was found
            on the suite variable ${WORKBOOK}.

        Examples:
        | Save excel changes | reinf | R2040.xlsx |

        =>

        | value |
        """
        wb = self.builtin.get_variable_value("${WORKBOOK}", work_book)
        sheet = wb.get_sheet_by_name(sheet)
        value = sheet.cell(row=row, column=column).value
        return value
