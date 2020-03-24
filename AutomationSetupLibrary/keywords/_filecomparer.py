import difflib
import sys
import os
import re
from robot.api import logger
from robot.libraries.BuiltIn import BuiltIn
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from jsondiff import diff
from .keywordgroup import KeywordGroup

class _FileComparer(KeywordGroup):

    def __init__(self):
        self.builtin = BuiltIn()

    def __compare_lines_difflib(self, benchLines, actualLines, outName, actualName):
        diffFiles = difflib.unified_diff(
            actualLines,
            benchLines
        )
        path, name = self.__split_file_name(actualName)
        fileName = outName + '.diff'
        curd = os.getcwd()
        os.chdir(path)
        sys.stdout.write(fileName)
        with open(fileName, 'w') as file_out:
            for line in diffFiles:
                file_out.write(line)
        file_out.close
        os.chdir(curd)
        return True

    def __get_file_name(self, fullFileName):
        #It is to reverse a string using extended slice syntax. string[begin:end:step]
        #e.g. fullFileName = ../folder/fileName.xml --> lmx.emaNelif/redlof/..
        reverse = fullFileName[::-1]
        #It is to get only the file name. e.g. lmx.emaNelif
        name = reverse[0:reverse.index("\\"):1]
        #Return file name --> fileName.xml
        return name[::-1]

    def __is_none(self, value):
        if value is None:
            return "None"
        else:
            return value

    def __split_file_name(self, fullFilename):
        path, fileName = os.path.split(fullFilename)
        return path, fileName

    def __open_xlsx_file(self, path, fileName):
        os.chdir(path)
        workbook = openpyxl.load_workbook(fileName)
        return workbook

    def compare_text_files(self, bench_filename, actual_filename):
        """Compares two text files.
        
        bench_filename and actual_filename are required. Fails if there are difference between files passed as arguments.
        
        Example:
        | Compare Two Text File | ${CURDIR}/bench.txt | ${CURDIR}/actual.txt|
        """
        identical = False
        with open(bench_filename, 'r') as benchFile:
            with open(actual_filename, 'r') as actualFile:
                diff = difflib.unified_diff(
                    actualFile.readlines(),
                    benchFile.readlines(),
                    fromfile='actualFile',
                    tofile='benchFile'
                )
                length = 0
                for line in diff:
                    length += 1
                    sys.stdout.write(line)
                if length == 0:
                    identical = True
            benchFile.close
        actualFile.close
        if identical == False:
            self.builtin.fail("Files are not equal.")

    def compare_text_files_using_regex(self, bench_filename, actual_filename):
        """Compares two text files using regular expression. It is useful to compare dynamic information like dates or numbers generated during execution.
        
        bench_filename and actual_filename are required. Fails if there are difference between files passed as arguments. If the comparison fails, it will
        generate a new file with the differences. 
        
        Example:
        Compare Text Files Using Regex | ${CURDIR}\\bench.xml | ${CURDIR}\\actual.xml|
        """
        identical = False
        diffName = self.__get_file_name(bench_filename)
        with open(bench_filename, 'r') as benchFile:
            with open(actual_filename, 'r') as actualFile:
                actual = actualFile.readlines()
                bench = benchFile.readlines()
                if len(bench) == len(actual):
                    rowDiff = 0
                    for benchLine, actualLine in zip(bench, actual):
                        benchLineC = benchLine.strip()
                        actualLineC = actualLine.strip()
                        if benchLineC != actualLineC:
                            pattern = re.compile(benchLineC)
                            if pattern.match(actualLineC) == None:
                                rowDiff += 1
                                msjError = "Actual line {} is not equal to expected {} .\r\n".format(actualLineC, benchLineC)
                                sys.stdout.write(msjError)
                    if rowDiff == 0:
                        identical = True
                else:
                    msjError = "Files have different number of lines: Actual file has {} lines but Benchmark file has {} lines.\r\n".format(len(actual), len(bench))
                    sys.stdout.write(msjError)
                if identical == False:
                    self.__compare_lines_difflib(bench, actual, diffName, actual_filename)
            benchFile.close
        actualFile.close
        if identical == False:
              self.builtin.fail("Files are not equal. Please review output file generated {}.diff".format(diffName))

    def compare_xlsx_files(self, bench_filename, actualfilename):
        """Compares two excel files and generate a new one with the differences.

        bench_filename and actual_filename are required. Fails if there are difference between files passed as arguments. If the comparison fails, it will
        generate a new file with the differences. 

        Example:
        | Compare Xlsx Files | ${CURDIR}\\bench.xlsx | ${CURDIR}\\actual.xlsx|
        """
        redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
        identical = False
        rowDiff = 0
        pathBench, nameBench = self.__split_file_name(bench_filename)
        pathActual, nameActual = self.__split_file_name(actualfilename)
        generalcwd = os.getcwd()
        benchWB =  self.__open_xlsx_file(pathBench, nameBench)
        actualWB =  self.__open_xlsx_file(pathActual, nameActual)
        os.chdir(generalcwd)
        name, extension = nameBench.split(".")
        diffName = "{}_diff.{}".format(name, extension)
        #create a workbook to save differences
        diffWB = Workbook()
        ws1 = diffWB.active
        ws1.title = "Differences"
        sheetsBench = benchWB.get_sheet_names()
        sheetsActual = actualWB.get_sheet_names()
        #checking if both files have the same sheets
        if sheetsBench == sheetsActual:
            for i in range(0, len(sheetsBench)):
                sheetBen = benchWB.get_sheet_by_name(sheetsBench[i])
                sheetAct = actualWB.get_sheet_by_name(sheetsBench[i])
                #checking if files have the same number of rows/columns
                if sheetBen.dimensions == sheetAct.dimensions:
                    max_rowB = sheetBen.max_row + 1
                    max_colB = sheetBen.max_column + 1
                    #comparing cell by cell
                    for row in range(1, max_rowB):
                        for column in range(1, max_colB):
                            cellBench = self.__is_none(sheetBen.cell(column=column, row=row).value)
                            cellActual = self.__is_none(sheetAct.cell(column=column, row=row).value)
                            if cellBench != cellActual:
                                rowDiff += 1
                                difference = "{}".format(cellBench) + " ---> " + "{}".format(cellActual)
                                ws1.cell(column=column, row=row, value=difference).fill = redFill
                            else:
                                ws1.cell(column=column, row=row, value=sheetBen.cell(row=row, column=column).value)
                    if rowDiff != 0:
                        print("There are differences between files. Please check output generated.")
                        print("before save diffname {}".format(diffName))
                        diffWB.save(filename = "{}".format(diffName))
                    else:
                        identical = True
                else:
                    print("Files have different dimensions. Please review generated file.")
        else:
            print("Files have different number of sheets. Please review generated file.")
        fullname = os.path.join(generalcwd, diffName)
        if identical == False:
              self.builtin.fail("Files are not equal. Please review output file generated {}".format(diffName))

    def compare_json_object(self, bench_json, actual_json):
        """Compares two json objects.

        bench_json and actual_json are required. Fails if there are difference between objects passed as arguments.

        Example:
        | Compare Json Objects | ${bench} | ${actual}|
        """
        identical = False
        result = diff(bench_json, actual_json)
        if result != {}:
            self.builtin.fail("Actual Value is not equal as expected. Differences found in the actual result were {}".format(result))
